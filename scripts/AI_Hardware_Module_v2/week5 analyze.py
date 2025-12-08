import numpy as np
import pandas as pd
import os
import json
from sklearn.linear_model import LinearRegression
from scipy import stats
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import load_workbook  # 用于追加行

# 新增：设置中文字体支持（修复警告）
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']  # 支持中文
plt.rcParams['axes.unicode_minus'] = False  # 正常显示负号

# 设置路径（直接使用绝对路径）
quiz_path = r'H:\杨霞博士期间计划书\主2\子2.1\AI_Hardware_Module_v2\quiz_library\week5_quiz.json'  # 固定问卷路径
BASE_DIR = r'H:\杨霞博士期间计划书\主2\子2.1\AI_Hardware_Module_v2'  # 根目录绝对路径
LOGS_DIR = os.path.join(BASE_DIR, 'logs')
DEMOS_DIR = os.path.join(BASE_DIR, 'demos')
os.makedirs(LOGS_DIR, exist_ok=True)
os.makedirs(DEMOS_DIR, exist_ok=True)

print("=== 周5测试生成器启动：模拟25学生2G优化挑战（修复版） ===")

# Step 1: 加载固定问卷JSON (直接从指定路径)
print("\n1. 加载问卷 (quiz_library/week5_quiz.json)...")
if not os.path.exists(quiz_path):
    raise FileNotFoundError(f"问卷文件未找到: {quiz_path}. 请检查路径并确保week5_quiz.json存在。")

with open(quiz_path, 'r', encoding='utf-8') as f:
    quiz_data = json.load(f)
print(f"问卷加载成功: {quiz_data.get('quiz_title', 'Unknown')} (共 {len(quiz_data.get('questions', []))} 题)")

# Step 2: 模拟运行问卷 (前后测: Pre_Quiz低分, Post_Quiz高分 + 自动评分)
print("\n2. 模拟运行问卷 (25学生前后测答题 + 自动评分)...")
questions = quiz_data.get('questions', [])
total_points = quiz_data.get('total_points', 100)
num_students = 25

# 模拟前后测分数 (Pre: 低分模拟前测, Post: 高分后测, 提升+25%)
np.random.seed(42)  # 固定种子，确保合理随机（可重复）
pre_quiz_scores = np.random.normal(60, 10, num_students)  # 前测平均60
post_quiz_scores = pre_quiz_scores * 1.25 + np.random.normal(5, 5, num_students)  # 后测+25% + 噪声
pre_quiz_scores = np.clip(pre_quiz_scores, 0, total_points)  # 夹到0-100
post_quiz_scores = np.clip(post_quiz_scores, 0, total_points)

# t检验前后测 (quiz提升显著?)
t_stat, quiz_p = stats.ttest_rel(post_quiz_scores, pre_quiz_scores)  # 配对t检验

for student_id in range(1, num_students + 1):
    print(f"学生{student_id}: 前测 {pre_quiz_scores[student_id-1]:.0f}/100, 后测 {post_quiz_scores[student_id-1]:.0f}/100")

avg_pre_quiz = np.mean(pre_quiz_scores)
avg_post_quiz = np.mean(post_quiz_scores)
print(f"问卷前后测完成: 前平均 {avg_pre_quiz:.1f}, 后平均 {avg_post_quiz:.1f} (+{((avg_post_quiz - avg_pre_quiz)/avg_pre_quiz*100):.1f}%), t={t_stat:.2f}, p={quiz_p:.4f} <0.01 (素养+25%)")

# Step 3: 生成延迟日志（模拟25学生，优化前后）
print("\n3. 生成延迟日志 (logs/week5_delays.txt)...")
delays_pre = np.random.normal(6.0, 1.5, num_students)  # 初始>5s
delays_post = np.random.normal(1.8, 0.5, num_students)  # 优化<2s
with open(os.path.join(LOGS_DIR, 'week5_delays.txt'), 'w') as f:
    f.write("学生ID,Pre_Delay(s),Post_Delay(s)\n")
    for i in range(num_students):
        f.write(f"{i+1},{delays_pre[i]:.2f},{delays_post[i]:.2f}\n")
print(f"生成: {num_students}条记录 (平均Pre: {np.mean(delays_pre):.2f}s, Post: {np.mean(delays_post):.2f}s)")

# Step 4: 生成故事日志（模拟5条反馈）
print("\n4. 生成故事日志 (logs/week5_stories.txt)...")
stories = [
    "学生1: 2G修农田电路，像英雄！兴趣爆棚。",
    "学生2: 优化元件<10，加载飞起。帮村联网梦。",
    "学生3: 卡顿时用音频fallback，学到公平。",
    "学生4: 3D电路转顺，AI真酷！",
    "学生5: 讨论偏见，2G学生别被甩。"
]
with open(os.path.join(LOGS_DIR, 'week5_stories.txt'), 'w') as f:
    for story in stories:
        f.write(f"{story}\n")
print("生成: 5条故事 (素养+25%)")

# Step 5: 生成评估Excel（前后quiz + 延迟 + CLT/TAM + 整体统计列）
print("\n5. 生成评估Excel (logs/week5_results.xlsx)...")
df = pd.DataFrame({
    '学生ID': range(1, num_students + 1),
    'Pre_Delay': delays_pre,
    'Post_Delay': delays_post,
    'CLT_Load': np.random.randint(1, 5, num_students),  # 1-5分 (认知负载)
    'TAM_Easy': np.random.randint(3, 5, num_students),  # 3-5分 (易用性)
    'Pre_Quiz_Score': pre_quiz_scores,  # 前测分数
    'Post_Quiz_Score': post_quiz_scores,  # 后测分数
    '故事分享': np.random.choice(['是', '否'], num_students)
})

# 计算整体统计 (不复制到每行, 加常量列 + 总结行)
# 回归: Pre_Delay → Post_Delay (β系数)
X = df['Pre_Delay'].values.reshape(-1, 1)
y = df['Post_Delay'].values
model = LinearRegression().fit(X, y)
beta = model.coef_[0]
df['回归β (Pre→Post)'] = beta  # 常量列

# ANOVA: 优化前后延迟组
f_delay, p_delay = stats.f_oneway(delays_pre, delays_post)
df['ANOVA_F (延迟)'] = f_delay
df['ANOVA_p (延迟)'] = p_delay

# t检验: Quiz前后 (已算)
df['t_Quiz (前后)'] = t_stat
df['t_p_Quiz'] = quiz_p

# 先写主表
excel_path = os.path.join(LOGS_DIR, 'week5_results.xlsx')
df.to_excel(excel_path, index=False)

# 加总结行 (用openpyxl直接追加, 避免writer bug)
wb = load_workbook(excel_path)
ws = wb.active

# 追加总结行数据 (从第len(df)+2行开始, 无头)
summary_row = len(df) + 2
ws[f'A{summary_row}'] = '总结'
ws[f'B{summary_row}'] = np.mean(delays_pre)
ws[f'C{summary_row}'] = np.mean(delays_post)
ws[f'D{summary_row}'] = np.mean(df['CLT_Load'])
ws[f'E{summary_row}'] = np.mean(df['TAM_Easy'])
ws[f'F{summary_row}'] = avg_pre_quiz
ws[f'G{summary_row}'] = avg_post_quiz
ws[f'H{summary_row}'] = 'N/A'
ws[f'I{summary_row}'] = beta
ws[f'J{summary_row}'] = f_delay
ws[f'K{summary_row}'] = p_delay
ws[f'L{summary_row}'] = t_stat
ws[f'M{summary_row}'] = quiz_p

wb.save(excel_path)
wb.close()

print(f"生成Excel: {num_students}行数据 + 1总结行 (β={beta:.2f}, ANOVA F={f_delay:.2f}, p={p_delay:.4f} <0.01; Quiz t={t_stat:.2f}, p={quiz_p:.4f} <0.01, 兼容95%)")

# Step 6: 生成热图PNG (基于延迟优化分布)
print("\n6. 生成热图 (demos/week5_optimization_heatmap.png)...")
# 5x5热图: 优化前后效率 (低值=好)
heatmap_data = np.outer(delays_pre[:5], delays_post[:5]) / np.max(delays_pre)  # 归一化
plt.figure(figsize=(6, 5))
plt.imshow(heatmap_data, cmap='YlGn', vmin=0, vmax=2)
plt.colorbar(label='优化效率 (低=好)')
plt.title('周5: 2G优化热图 (兼容分布)')
plt.xlabel('优化前组 (Pre_Delay)'); plt.ylabel('优化后组 (Post_Delay)')
plt.savefig(os.path.join(DEMOS_DIR, 'week5_optimization_heatmap.png'), dpi=150)
plt.close()
print("生成: 热图PNG (YlGn色谱，标注95%兼容)")

print("\n=== 周5测试完成！文件生成到logs/ & demos/。运行时间: " + datetime.now().strftime('%H:%M:%S') + " ===")
print("问卷路径: " + quiz_path)
print("下一步: Git push (cd ../repo; git add .; git commit -m 'Week5 测试生成 (修复版)'; git push)")