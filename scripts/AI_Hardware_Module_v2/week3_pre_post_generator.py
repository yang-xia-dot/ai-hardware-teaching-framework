import openpyxl
import os
import random  # 用于模拟虚拟学生数据

# Define the file path (same as previous planning)
file_path = r'H:\杨霞博士期间计划书\主2\子2.1\AI_Hardware_Module_v2\logs\week3_pre_post.xlsx'

# Create directory if not exists
os.makedirs(os.path.dirname(file_path), exist_ok=True)

# Create a new workbook
wb = openpyxl.Workbook()

# Sheet 1: Data (学生数据，包括模拟的前测和后测分数)
ws_data = wb.active
ws_data.title = 'Data'

# Define columns: Student ID, Pre-Test Total Score, Post-Test Total Score, Q1 Pre, Q1 Post, ..., Q10 Pre, Q10 Post
columns = ['Student ID', 'Pre-Test Total Score', 'Post-Test Total Score']
for i in range(1, 11):
    columns.append(f'Q{i} Pre')
    columns.append(f'Q{i} Post')

# Write the header row
for col_num, column_title in enumerate(columns, 1):
    ws_data.cell(row=1, column=col_num, value=column_title)

# Define correct answers for objective questions (for simulation)
correct_answers = {
    'Q1': 'A',
    'Q2': 'A',
    'Q3': 'B',
    'Q4': 'B',
    'Q6': 'B',
    'Q7': 'A'
}

# Simulate data for 25 students (S001 to S025)
for row in range(2, 27):  # Rows 2 to 26
    student_id = f'S{row - 1:03d}'
    ws_data.cell(row=row, column=1, value=student_id)

    pre_scores = []
    post_scores = []

    for q in range(1, 11):
        q_str = f'Q{q}'
        if q_str in correct_answers:  # Objective questions
            # 前测: 低正确率 (~40%)
            pre_correct = random.random() < 0.4
            pre_score = 5 if pre_correct else 0
            # 后测: 高正确率 (~80%)
            post_correct = random.random() < 0.8
            post_score = 5 if post_correct else 0
        else:  # Self-assessment questions (Q5, Q8-Q10)
            # 前测: 低分 (1-3, 平均~2)
            pre_score = random.randint(1, 3)
            # 后测: 高分 (3-5, 平均~4)
            post_score = random.randint(3, 5)

        pre_scores.append(pre_score)
        post_scores.append(post_score)

        # Write scores to columns (Q1 Pre is column 4, Q1 Post column 5, etc.)
        pre_col = 3 + (q - 1) * 2 + 1  # Starting from column 4, step 2
        post_col = pre_col + 1
        ws_data.cell(row=row, column=pre_col, value=pre_score)
        ws_data.cell(row=row, column=post_col, value=post_score)

    # Calculate and write total scores
    pre_total = sum(pre_scores)
    post_total = sum(post_scores)
    ws_data.cell(row=row, column=2, value=pre_total)
    ws_data.cell(row=row, column=3, value=post_total)

# Sheet 2: Question Template (问题模板，像之前一样)
ws_template = wb.create_sheet(title='Question Template')

# Write question template (same as before, with your specified format)
template_content = [
    ['🔵 第二部分：操作系统知识测试 | OS Knowledge Assessment', '', ''],
    ['（4题，每题5分，共20分）', '', ''],
    ['Q1. 操作系统安装的分区步骤正确顺序是哪几个？（单选）', '', '（教师评分用：正确答案A，5分）'],
    ['What is the correct order of partitioning steps in OS installation? (Single choice)', '', ''],
    ['', '☐ A. 下载镜像 → 分区配置 → 引导启动', ''],
    ['', '      Download image → Partition setup → Boot setup', ''],
    ['', '☐ B. 引导启动 → 分区配置 → 下载镜像', ''],
    ['', '      Boot setup → Partition setup → Download image', ''],
    ['', '☐ C. 分区配置 → 引导启动 → 安装驱动', ''],
    ['', '      Partition setup → Boot setup → Install drivers', ''],
    ['', '☐ D. 安装驱动 → 下载镜像 → 引导启动', ''],
    ['', '      Install drivers → Download image → Boot setup', ''],
    ['Q2. 驱动程序的主要作用是什么？（单选）', '', '（教师评分用：正确答案A，5分）'],
    ['What is the primary role of drivers? (Single choice)', '', ''],
    ['', '☐ A. 管理硬件和软件的桥接', ''],
    ['', '      Bridge hardware and software', ''],
    ['', '☐ B. 增加操作系统速度', ''],
    ['', '      Increase OS speed', ''],
    ['', '☐ C. 替代操作系统', ''],
    ['', '      Replace the operating system', ''],
    ['', '☐ D. 仅用于量子芯片', ''],
    ['', '      Only for quantum chips', ''],
    ['Q3. 在QEMU模拟安装中，如果遇到兼容性问题，应如何处理？（单选）', '', '（教师评分用：正确答案B，5分）'],
    ['How to handle compatibility issues in QEMU simulation? (Single choice)', '', ''],
    ['', '☐ A. 更换硬件', ''],
    ['', '      Replace hardware', ''],
    ['', '☐ B. 检查驱动版本或调整分区', ''],
    ['', '      Check driver version or adjust partitioning', ''],
    ['', '☐ C. 忽略问题继续安装', ''],
    ['', '      Ignore and proceed', ''],
    ['', '☐ D. 关闭虚拟机', ''],
    ['', '      Shut down the virtual machine', ''],
    ['Q4. 从真空管到量子芯片，OS安装的要求如何变化？（单选）', '', '（教师评分用：正确答案B，5分）'],
    ['How have OS installation requirements changed from vacuum tubes to quantum chips? (Single choice)', '', ''],
    ['', '☐ A. 更简单，无需分区', ''],
    ['', '      Simpler, no partitioning needed', ''],
    ['', '☐ B. 更复杂，需更高兼容性', ''],
    ['', '      More complex, higher compatibility required', ''],
    ['', '☐ C. 完全不变', ''],
    ['', '      No change at all', ''],
    ['', '☐ D. 仅依赖网速', ''],
    ['', '      Only dependent on internet speed', ''],
    ['🟢 第三部分：AI伦理意识评估 | AI Ethics Awareness Assessment', '', ''],
    ['（3题，每题5分，共15分）', '', ''],
    ['Q5. 安装软件过程中，隐私泄露的风险如何避免？（自评）', '', '（教师评分用：目标后测平均分≥3.5）'],
    ['How can privacy leakage risks be avoided during software installation? (Self-assessment)', '', ''],
    ['', '请根据你的了解程度打分（1-5分）：', ''],
    ['', 'Please rate your understanding level (1-5 scale):', ''],
    ['', '☐ 1分 - 完全不知道 Completely unaware', ''],
    ['', '☐ 2分 - 知道一点 Know a little', ''],
    ['', '☐ 3分 - 基本了解 Basically understand', ''],
    ['', '☐ 4分 - 比较了解 Fairly understand', ''],
    ['', '☐ 5分 - 完全了解 Fully understand', ''],
    ['', '请简述你的理解（选填）： Please briefly describe your understanding (optional):', ''],
    ['Q6. 如果乡村数据被AI收集，会如何影响公平性？（单选）', '', '（教师评分用：正确答案B，5分）'],
    ['How does AI collecting rural data affect fairness? (Single choice)', '', ''],
    ['', '☐ A. 提升城乡平等', ''],
    ['', '      Enhance urban-rural equality', ''],
    ['', '☐ B. 加剧城乡鸿沟', ''],
    ['', '      Widen urban-rural divide', ''],
    ['', '☐ C. 无任何影响', ''],
    ['', '      No impact', ''],
    ['', '☐ D. 仅改善城市', ''],
    ['', '      Only improves urban areas', ''],
    ['Q7. OS升级失败因数据偏见加剧城乡鸿沟的例子是？（单选）', '', '（教师评分用：正确答案A，5分）'],
    ['An example of OS upgrade failure due to data bias widening urban-rural divide is? (Single choice)', '', ''],
    ['', '☐ A. 城市算法优先，农村兼容性差', ''],
    ['', '      Urban algorithms prioritized, rural compatibility poor', ''],
    ['', '☐ B. 农村网速快于城市', ''],
    ['', '      Rural internet faster than urban', ''],
    ['', '☐ C. 量子芯片无偏见', ''],
    ['', '      Quantum chips have no bias', ''],
    ['', '☐ D. 数据泄露无影响', ''],
    ['', '      Data leaks have no impact', ''],
    ['🟡 第四部分：学习兴趣与参与度评估 | Interest & Engagement Assessment', '', ''],
    ['（3题，每题5分，共15分）', '', ''],
    ['Q8. 你对芯片从真空管到量子芯片的历史感兴趣吗？（自评）', '', '（教师评分用：目标后测平均分≥3.5）'],
    ['Are you interested in the history of chips from vacuum tubes to quantum chips? (Self-assessment)', '', ''],
    ['', '请根据你的兴趣程度打分（1-5分）：', ''],
    ['', 'Please rate your interest level (1-5 scale):', ''],
    ['', '☐ 1分 - 完全不感兴趣 Not interested at all', ''],
    ['', '☐ 2分 - 不太感兴趣 Not very interested', ''],
    ['', '☐ 3分 - 一般 Neutral', ''],
    ['', '☐ 4分 - 比较感兴趣 Fairly interested', ''],
    ['', '☐ 5分 - 非常感兴趣 Very interested', ''],
    ['Q9. 你觉得"OS升级剧本工坊"活动有趣吗？（自评）', '', '（教师评分用：目标后测平均分≥3.5）'],
    ['Do you find the "OS Upgrade Script Workshop" activity interesting? (Self-assessment)', '', ''],
    ['', '请根据你的真实感受打分（1-5分）：', ''],
    ['', 'Please rate your true feelings (1-5 scale):', ''],
    ['', '☐ 1分 - 完全无趣 Not fun at all', ''],
    ['', '☐ 2分 - 不太有趣 Not very fun', ''],
    ['', '☐ 3分 - 一般 Neutral', ''],
    ['', '☐ 4分 - 比较有趣 Fairly fun', ''],
    ['', '☐ 5分 - 非常有趣 Very fun', ''],
    ['Q10. 你愿意参与下次跨学科技术活动吗？（自评）', '', '（教师评分用：目标后测平均分≥3.5）'],
    ['Would you participate in the next cross-disciplinary tech activity? (Self-assessment)', '', ''],
    ['', '请根据你的真实意愿打分（1-5分）：', ''],
    ['', 'Please rate your true willingness (1-5 scale):', ''],
    ['', '☐ 1分 - 完全不愿意 Completely unwilling', ''],
    ['', '☐ 2分 - 不太愿意 Not very willing', ''],
    ['', '☐ 3分 - 无所谓 Indifferent', ''],
    ['', '☐ 4分 - 比较愿意 Fairly willing', ''],
    ['', '☐ 5分 - 非常愿意 Very willing', ''],
    ['📊 第五部分：开放式反思（选填）| Open-ended Reflection (Optional)', '', ''],
    ['课后反思（仅后测填写）| Post-class Reflection (Post-test only)', '', ''],
    ['', '1. 这节课你最大的收获是什么？ What is your biggest takeaway from this class?', ''],
    ['', '2. 你在学习中遇到的困难是什么？ What difficulties did you encounter during learning?', ''],
    ['', '3. 你对下节课（周4视觉AI实战）有什么期待？ What are your expectations for the next class (Week 4 Visual AI)?',
     ''],
    ['✅ 问卷完成确认 | Survey Completion Confirmation', '', ''],
    ['', '☐ 我已完成全部10题问卷填写 I have completed all 10 questions', ''],
    ['', '☐ 我的答案真实反映我的理解和感受 My answers truly reflect my understanding and feelings', ''],
    ['', '学生签名 / Student Signature: _', ''],
    ['', '提交时间 / Submission Time: _:_', ''],
    ['📋 教师评分表 | Teacher Scoring Sheet', '', ''],
    ['', '（教师使用，学生无需填写）', ''],
    ['题号 正确答案/目标分 学生得分 备注', '', ''],
    ['Q1 A (5分)', '', ''],
    ['Q2 A (5分)', '', ''],
    ['Q3 B (5分)', '', ''],
    ['Q4 B (5分)', '', ''],
    ['Q5 目标≥3.5分', '', ''],
    ['Q6 B (5分)', '', ''],
    ['Q7 A (5分)', '', ''],
    ['Q8 目标≥3.5分', '', ''],
    ['Q9 目标≥3.5分', '', ''],
    ['Q10 目标≥3.5分', '', ''],
    ['总分 / Total 50分', '', ''],
    ['前后测对比分析 / Pre-Post Comparison:', '', ''],
    ['', '* 前测平均分 Pre-test Average: 分 (基准 Baseline: <30分)', ''],
    ['', '* 后测平均分 Post-test Average: 分 (目标 Target: ≥36分, +20%)', ''],
    ['', '* 提升幅度 Improvement: % (目标 Target: ≥20%)', ''],
    ['', '教师签名 / Teacher Signature: _', ''],
    ['', '评分日期 / Scoring Date: _年_月_日', '']
]

# Write the template content to the sheet
for row_num, row_data in enumerate(template_content, 1):
    for col_num, value in enumerate(row_data, 1):
        ws_template.cell(row=row_num, column=col_num, value=value)

# Sheet 3: Scoring Sheet (教师评分表，自动计算平均)
ws_scoring = wb.create_sheet(title='Scoring Sheet')

# Header for scoring
scoring_headers = ['Question', 'Correct Answer/Target', 'Pre Avg Score', 'Post Avg Score', 'Remarks']
for col_num, header in enumerate(scoring_headers, 1):
    ws_scoring.cell(row=1, column=col_num, value=header)

# Calculate averages from Data sheet
for q in range(1, 11):
    row_num = q + 1
    q_str = f'Q{q}'
    pre_col = 3 + (q - 1) * 2 + 1  # Pre column in Data
    post_col = pre_col + 1  # Post column in Data

    # Calculate pre avg
    pre_sum = sum([ws_data.cell(row=r, column=pre_col).value for r in range(2, 27) if
                   ws_data.cell(row=r, column=pre_col).value is not None])
    pre_avg = pre_sum / 25

    # Calculate post avg
    post_sum = sum([ws_data.cell(row=r, column=post_col).value for r in range(2, 27) if
                    ws_data.cell(row=r, column=post_col).value is not None])
    post_avg = post_sum / 25

    correct = correct_answers.get(q_str, '目标≥3.5分')
    ws_scoring.cell(row=row_num, column=1, value=q_str)
    ws_scoring.cell(row=row_num, column=2, value=correct)
    ws_scoring.cell(row=row_num, column=3, value=pre_avg)
    ws_scoring.cell(row=row_num, column=4, value=post_avg)
    ws_scoring.cell(row=row_num, column=5, value='')

# Total averages
pre_total_avg = sum([ws_data.cell(row=r, column=2).value for r in range(2, 27)]) / 25
post_total_avg = sum([ws_data.cell(row=r, column=3).value for r in range(2, 27)]) / 25
improvement = ((post_total_avg - pre_total_avg) / pre_total_avg) * 100 if pre_total_avg > 0 else 0

ws_scoring.cell(row=12, column=1, value='Total')
ws_scoring.cell(row=12, column=2, value='50分')
ws_scoring.cell(row=12, column=3, value=pre_total_avg)
ws_scoring.cell(row=12, column=4, value=post_total_avg)
ws_scoring.cell(row=13, column=1, value='前后测对比分析 / Pre-Post Comparison')
ws_scoring.cell(row=14, column=1, value='* 前测平均分 Pre-test Average')
ws_scoring.cell(row=14, column=2, value=f'{pre_total_avg:.2f} 分 (基准 Baseline: <30分)')
ws_scoring.cell(row=15, column=1, value='* 后测平均分 Post-test Average')
ws_scoring.cell(row=15, column=2, value=f'{post_total_avg:.2f} 分 (目标 Target: ≥36分, +20%)')
ws_scoring.cell(row=16, column=1, value='* 提升幅度 Improvement')
ws_scoring.cell(row=16, column=2, value=f'{improvement:.2f}% (目标 Target: ≥20%)')
ws_scoring.cell(row=17, column=1, value='教师签名 / Teacher Signature: _')
ws_scoring.cell(row=18, column=1, value='评分日期 / Scoring Date: _年_月_日')

# Save the workbook
wb.save(file_path)
print(f'Excel file generated and saved at: {file_path}')
print(f'模拟前测平均分: {pre_total_avg:.2f}, 后测平均分: {post_total_avg:.2f}, 提升: {improvement:.2f}%')